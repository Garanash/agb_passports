"""
Генератор наклеек в Excel формате из шаблона.
Все наклейки в один столбец. Колонки A=100px, B=200px, C=100px. Логотип 1.07×3.61 см.
"""
import os
import io
import sys
from typing import List
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.worksheet.pagebreak import Break
from backend.utils.template_manager import get_template_manager
from backend.utils.barcode_generator import generate_barcode_image


def copy_template_to_target(template_ws, target_ws, row_offset, col_offset):
    """
    Полностью копирует шаблон в целевой лист начиная с указанной позиции
    Копирует: значения, стили, размеры, объединения
    """
    STICKER_ROWS = template_ws.max_row
    STICKER_COLS = template_ws.max_column
    
    # 1. Копируем размеры колонок
    for col_idx in range(1, STICKER_COLS + 1):
        col_letter = get_column_letter(col_idx)
        source_width = template_ws.column_dimensions[col_letter].width
        if source_width:
            target_col_letter = get_column_letter(col_offset + col_idx - 1)
            target_ws.column_dimensions[target_col_letter].width = source_width
    
    # 2. Копируем высоты строк
    for row_idx in range(1, STICKER_ROWS + 1):
        source_height = template_ws.row_dimensions[row_idx].height
        if source_height:
            target_row = row_offset + row_idx - 1
            target_ws.row_dimensions[target_row].height = source_height
    
    # 3. Копируем объединенные ячейки
    for merged_range in template_ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        target_min_col = col_offset + min_col - 1
        target_min_row = row_offset + min_row - 1
        target_max_col = col_offset + max_col - 1
        target_max_row = row_offset + max_row - 1
        
        target_ws.merge_cells(
            start_row=target_min_row,
            start_column=target_min_col,
            end_row=target_max_row,
            end_column=target_max_col
        )
    
    # 4. Копируем все ячейки со значениями и стилями
    for row_idx in range(1, STICKER_ROWS + 1):
        for col_idx in range(1, STICKER_COLS + 1):
            source_cell = template_ws.cell(row_idx, col_idx)
            target_row = row_offset + row_idx - 1
            target_col = col_offset + col_idx - 1
            target_cell = target_ws.cell(target_row, target_col)
            
            # Проверяем, не является ли ячейка частью объединения (MergedCell)
            # Если да, то значение копируем только из главной ячейки
            try:
                # Проверяем, является ли target_cell объединенной ячейкой
                # После merge_cells, все ячейки кроме первой становятся MergedCell
                if isinstance(target_cell, MergedCell):
                    # Это объединенная ячейка, пропускаем копирование значения
                    # Значение будет в главной ячейке объединения
                    pass
                else:
                    # Это обычная ячейка, можно копировать значение
                    if source_cell.value is not None:
                        target_cell.value = source_cell.value
            except (AttributeError, TypeError) as e:
                # Если возникла ошибка, пропускаем эту ячейку
                pass
            
            # Копируем стили (только для обычных ячеек, не для MergedCell)
            # Проверяем, не является ли ячейка объединенной
            if isinstance(target_cell, MergedCell):
                # Пропускаем копирование стилей для объединенных ячеек
                pass
            else:
                try:
                    if source_cell.font:
                        try:
                            target_cell.font = Font(
                                name=source_cell.font.name,
                                size=source_cell.font.size,
                                bold=source_cell.font.bold,
                                italic=source_cell.font.italic,
                                color=source_cell.font.color
                            )
                        except:
                            pass
                    
                    if source_cell.alignment:
                        try:
                            target_cell.alignment = Alignment(
                                horizontal=source_cell.alignment.horizontal,
                                vertical=source_cell.alignment.vertical,
                                wrap_text=source_cell.alignment.wrap_text,
                                indent=getattr(source_cell.alignment, 'indent', 0),
                                text_rotation=getattr(source_cell.alignment, 'text_rotation', 0)
                            )
                        except:
                            pass
                    
                    if source_cell.border:
                        try:
                            source_border = source_cell.border
                            target_cell.border = Border(
                                left=Side(style=source_border.left.style, color=source_border.left.color) if source_border.left and source_border.left.style else None,
                                right=Side(style=source_border.right.style, color=source_border.right.color) if source_border.right and source_border.right.style else None,
                                top=Side(style=source_border.top.style, color=source_border.top.color) if source_border.top and source_border.top.style else None,
                                bottom=Side(style=source_border.bottom.style, color=source_border.bottom.color) if source_border.bottom and source_border.bottom.style else None
                            )
                        except:
                            pass
                    
                    if source_cell.fill:
                        try:
                            source_fill = source_cell.fill
                            if hasattr(source_fill, 'patternType'):
                                target_cell.fill = PatternFill(
                                    patternType=source_fill.patternType,
                                    fgColor=source_fill.fgColor,
                                    bgColor=getattr(source_fill, 'bgColor', None)
                                )
                        except:
                            pass
                except (AttributeError, TypeError):
                    pass


def find_and_replace_cell(ws, row_offset, col_offset, sticker_rows, sticker_cols, search_text, new_value):
    """Находит ячейку по тексту или переменной и заменяет значение"""
    for row_idx in range(row_offset, row_offset + sticker_rows):
        for col_idx in range(col_offset, col_offset + sticker_cols):
            cell = ws.cell(row_idx, col_idx)
            if cell.value and isinstance(cell.value, str):
                # Проверяем, содержит ли ячейка искомый текст или переменную
                if search_text in cell.value:
                    # Заменяем весь текст или только переменную
                    if '{{' in cell.value and '}}' in cell.value:
                        # Если есть переменные в формате {{ variable }}, заменяем их
                        import re
                        # Ищем переменную в формате {{ search_text }} или просто search_text
                        pattern = r'\{\{\s*' + re.escape(search_text.replace('{{', '').replace('}}', '').strip()) + r'\s*\}\}'
                        if re.search(pattern, cell.value, re.IGNORECASE):
                            cell.value = re.sub(pattern, str(new_value), cell.value, flags=re.IGNORECASE)
                        else:
                            cell.value = new_value
                    else:
                        cell.value = new_value
                    return row_idx, col_idx
    return None, None


def _get_image_cell_position(image):
    """Возвращает (row_1based, col_1based) для изображения или (None, None). openpyxl anchor — 0-based."""
    try:
        if not hasattr(image, 'anchor') or not image.anchor:
            return None, None
        anchor = image.anchor
        row0, col0 = None, None
        if hasattr(anchor, '_from') and anchor._from:
            row0 = getattr(anchor._from, 'row', None)
            col0 = getattr(anchor._from, 'col', None)
        if row0 is None or col0 is None:
            return None, None
        return int(row0) + 1, int(col0) + 1
    except Exception:
        return None, None


def _remove_images_in_region(ws, row_offset, col_offset, sticker_rows, sticker_cols):
    """Удаляет все изображения, чья верхняя левая ячейка попадает в область наклейки. Возвращает количество удалённых."""
    removed = 0
    if not hasattr(ws, '_images') or not ws._images:
        return removed
    to_remove = []
    for image in list(ws._images):
        img_row, img_col = _get_image_cell_position(image)
        if img_row is None:
            continue
        if (row_offset <= img_row < row_offset + sticker_rows and
                col_offset <= img_col < col_offset + sticker_cols):
            to_remove.append(image)
    for img in to_remove:
        try:
            ws._images.remove(img)
            removed += 1
        except Exception:
            pass
    return removed


def _remove_images_in_cell(ws, row_1based, col_1based):
    """Удаляет все изображения, привязанные к данной ячейке (1-based). Гарантирует отсутствие дублей перед вставкой."""
    if not hasattr(ws, '_images') or not ws._images:
        return
    to_remove = []
    for image in list(ws._images):
        img_row, img_col = _get_image_cell_position(image)
        if img_row == row_1based and img_col == col_1based:
            to_remove.append(image)
    for img in to_remove:
        try:
            ws._images.remove(img)
        except Exception:
            pass


def _find_merged_bounds(ws, row_idx, col_idx):
    """Возвращает bounds (min_col, min_row, max_col, max_row) объединения, если точка внутри; иначе None."""
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if min_row <= row_idx <= max_row and min_col <= col_idx <= max_col:
            return min_col, min_row, max_col, max_row
    return None


def _col_width_px(ws, col_1based):
    # В openpyxl ширина колонки ~ в "символах"; приближенно 7px за единицу
    col_letter = get_column_letter(col_1based)
    w = ws.column_dimensions[col_letter].width
    if not w:
        w = 10
    return float(w) * 7.0


def _row_height_px(ws, row_1based):
    # Высота строки в points; 1pt = 96/72 px
    h_pt = ws.row_dimensions[row_1based].height
    if not h_pt:
        h_pt = 15  # дефолт Excel
    return float(h_pt) * (96.0 / 72.0)


def _add_image_centered(ws, img: OpenpyxlImage, row_1based: int, col_1based: int, top_offset_px=None):
    """
    Вставляет изображение в ячейку (или объединение). По умолчанию — по центру.
    top_offset_px: если задано, картинка выравнивается по горизонтали по центру, по вертикали — отступ от верха (для текста над картинкой).
    img.width/img.height должны быть заданы в пикселях.
    """
    merged = _find_merged_bounds(ws, row_1based, col_1based)
    if merged:
        min_col, min_row, max_col, max_row = merged
    else:
        min_col = max_col = col_1based
        min_row = max_row = row_1based

    cell_w_px = sum(_col_width_px(ws, c) for c in range(min_col, max_col + 1))
    cell_h_px = sum(_row_height_px(ws, r) for r in range(min_row, max_row + 1))

    img_w_px = float(getattr(img, "width", 0) or 0)
    img_h_px = float(getattr(img, "height", 0) or 0)

    off_x_px = max((cell_w_px - img_w_px) / 2.0, 0.0)
    if top_offset_px is not None:
        off_y_px = max(float(top_offset_px), 0.0)
    else:
        off_y_px = max((cell_h_px - img_h_px) / 2.0, 0.0)

    EMU_PER_PX = 9525  # 1px = 9525 EMU
    marker = AnchorMarker(
        col=min_col - 1,
        colOff=int(off_x_px * EMU_PER_PX),
        row=min_row - 1,
        rowOff=int(off_y_px * EMU_PER_PX),
    )
    size = XDRPositiveSize2D(cx=int(img_w_px * EMU_PER_PX), cy=int(img_h_px * EMU_PER_PX))
    img.anchor = OneCellAnchor(_from=marker, ext=size)
    ws.add_image(img)


def _cell_display_value(ws, row_idx, col_idx):
    """Возвращает значение ячейки; для MergedCell — значение из верхней левой ячейки объединения."""
    cell = ws.cell(row_idx, col_idx)
    if not isinstance(cell, MergedCell):
        return cell.value
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if min_row <= row_idx <= max_row and min_col <= col_idx <= max_col:
            return ws.cell(min_row, min_col).value
    return None


def _find_cell_with_placeholder(ws, row_offset, col_offset, sticker_rows, sticker_cols, placeholder_name):
    """Ищет ячейку, содержащую {{ placeholder_name }}. Возвращает (row, col) или (None, None)."""
    import re
    pattern = r'\{\{\s*' + re.escape(placeholder_name) + r'\s*\}\}'
    for row_idx in range(row_offset, row_offset + sticker_rows):
        for col_idx in range(col_offset, col_offset + sticker_cols):
            cell = ws.cell(row_idx, col_idx)
            if isinstance(cell, MergedCell):
                continue
            if cell.value and isinstance(cell.value, str) and re.search(pattern, cell.value, re.IGNORECASE):
                return row_idx, col_idx
    return None, None


def replace_all_variables(ws, row_offset, col_offset, sticker_rows, sticker_cols, variables_dict):
    """Заменяет все переменные в формате {{ variable }} в области наклейки. Не трогает плейсхолдеры картинок (stock_code, serial_number_code)."""
    import re
    # Плейсхолдеры, которые подставляют картинки, не заменяем текстом
    image_placeholders = {'stock_code', 'serial_number_code'}
    for row_idx in range(row_offset, row_offset + sticker_rows):
        for col_idx in range(col_offset, col_offset + sticker_cols):
            cell = ws.cell(row_idx, col_idx)
            if isinstance(cell, MergedCell):
                continue
            if cell.value and isinstance(cell.value, str):
                original_value = cell.value
                for var_name, var_value in variables_dict.items():
                    if var_name in image_placeholders:
                        continue
                    pattern = r'\{\{\s*' + re.escape(var_name) + r'\s*\}\}'
                    if re.search(pattern, cell.value, re.IGNORECASE):
                        cell.value = re.sub(pattern, str(var_value), cell.value, flags=re.IGNORECASE)
                if cell.value != original_value:
                    print(f"    ✅ Заменена переменная в ячейке {get_column_letter(col_idx)}{row_idx}: {original_value} -> {cell.value[:50]}")


def generate_stickers_excel(passports, template_path=None):
    """
    Генерирует Excel файл с наклейками максимально точно как в шаблоне
    
    АЛГОРИТМ:
    1. Загружаем Excel шаблон
    2. Для каждой наклейки ПОЛНОСТЬЮ копируем шаблон (все ячейки, стили, границы, объединения)
    3. Заменяем только текстовые значения (переменные) данными из паспорта
    4. Добавляем/заменяем только логотип и штрихкоды
    5. Размещаем 4 наклейки на странице (2x2) БЕЗ отступов
    """
    print(f"📊 Генерация наклеек в Excel: {len(passports)} паспортов")
    sys.stdout.flush()
    
    # Получаем путь к шаблону
    if template_path is None:
        manager = get_template_manager()
        template_path_obj = manager.get_template_path("sticker")
        template_path = str(template_path_obj) if template_path_obj and template_path_obj.exists() else None
    
    # Проверяем, что шаблон Excel существует
    if not template_path or not os.path.exists(template_path) or not template_path.endswith('.xlsx'):
        raise ValueError(f"Excel шаблон не найден: {template_path}")
    
    try:
        # Загружаем Excel шаблон
        template_wb = load_workbook(template_path)
        template_ws = template_wb.active
        print(f"    ✅ Используется Excel шаблон: {template_path}")
        print(f"    📏 Размеры шаблона: {template_ws.max_row} строк × {template_ws.max_column} колонок")
    except Exception as e:
        print(f"    ❌ Не удалось загрузить Excel шаблон: {e}")
        import traceback
        traceback.print_exc()
        raise
    
    # Получаем путь к логотипу
    manager = get_template_manager()
    logo_path_obj = manager.get_logo_path()
    logo_path = str(logo_path_obj) if logo_path_obj and logo_path_obj.exists() else None
    
    if not logo_path:
        alt_paths = [
            '/app/templates/logo.png',
            'templates/logo.png'
        ]
        for alt_path in alt_paths:
            if os.path.exists(alt_path):
                logo_path = alt_path
                print(f"    ✅ Логотип найден по альтернативному пути: {logo_path}")
                break
    
    if not logo_path or not os.path.exists(logo_path):
        print(f"    ⚠️ Логотип не найден ни по одному из путей!")
    else:
        print(f"    ✅ Логотип найден: {logo_path}")
    
    # Размеры наклейки из шаблона
    STICKER_ROWS = template_ws.max_row
    STICKER_COLS = template_ws.max_column
    
    print(f"    📐 Размеры одной наклейки: {STICKER_ROWS} строк × {STICKER_COLS} колонок")
    
    # Создаем новую рабочую книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "Наклейки"
    
    # Список временных файлов для удаления после сохранения
    temp_files_to_cleanup = []
    
    # Все наклейки в один столбец: каждая следующая под предыдущей
    # Ширина колонок в пикселях: A=100, B=200, C=100 (Excel ~7 px на единицу ширины)
    COL_WIDTHS_PX = [100, 200, 100]
    # Логотип: размер 1.07 см x 3.61 см (1 см ≈ 37.8 px при 96 DPI)
    LOGO_WIDTH_CM, LOGO_HEIGHT_CM = 1.07, 3.61
    CM_TO_PX = 37.7952755906  # 96 DPI
    
    # Счётчик реально сгенерированных наклеек (нужен для аккуратного смещения и разрывов страниц)
    stickers_generated = 0
    
    for sticker_idx, passport in enumerate(passports):
        if not passport.nomenclature:
            continue
        
        # Смещение считаем по количеству уже добавленных наклеек, чтобы пропуски не ломали сетку
        row_offset = 1 + stickers_generated * STICKER_ROWS
        col_offset = 1
        
        print(f"    📍 Наклейка {stickers_generated + 1}: row_offset={row_offset}, col_offset={col_offset}")
        
        copy_template_to_target(template_ws, ws, row_offset, col_offset)
        # Фиксированная ширина колонок A, B, C в пикселях → единицы Excel
        for c in range(min(len(COL_WIDTHS_PX), STICKER_COLS)):
            _col_letter = get_column_letter(col_offset + c)
            ws.column_dimensions[_col_letter].width = COL_WIDTHS_PX[c] / 7.0
        print(f"    ✅ Шаблон скопирован для наклейки {sticker_idx + 1}")
        
        # Удаляем ВСЕ изображения в области этой наклейки
        images_removed_count = _remove_images_in_region(ws, row_offset, col_offset, STICKER_ROWS, STICKER_COLS)
        if images_removed_count > 0:
            print(f"    🗑️ Удалено {images_removed_count} изображений в области наклейки {sticker_idx + 1}")
        
        # Заполняем данными из паспорта
        nomenclature = passport.nomenclature
        
        # Получаем данные
        stock_code = nomenclature.article or getattr(nomenclature, 'code_1c', None) or '3501040'
        serial_number = passport.passport_number or 'AGB0000125'
        matrix_val = nomenclature.matrix or 'HQ'
        height_val = str(nomenclature.height or getattr(nomenclature, 'drilling_depth', None) or '12')
        waterways_val = str(getattr(nomenclature, 'waterways', None) or '8')
        
        # Формируем дату
        if passport.created_at:
            day = passport.created_at.strftime("%d")
            month = passport.created_at.strftime("%m")
            year = passport.created_at.strftime("%Y")
        else:
            from datetime import datetime
            now = datetime.now()
            day = now.strftime("%d")
            month = now.strftime("%m")
            year = now.strftime("%Y")
        
        # Заменяем переменные в шаблоне данными
        nom_name = nomenclature.name or 'Коронка импрегнированная'
        # Всегда после слова ALFA — перенос строки (как в шаблоне)
        if 'ALFA' in nom_name:
            nom_name = nom_name.replace('ALFA', 'ALFA\n', 1)
        
        # Словарь переменных для замены (в формате {{ variable }})
        variables = {
            'nomenclature_name': nom_name,
            'article': stock_code,
            'stock_code': stock_code,
            'matrix': matrix_val,
            'height': f"{height_val} мм.",
            'waterways': f"{waterways_val} мм.",
            'serial_number': serial_number,
            'serial number': serial_number,
            'production_date': f'«{day}» {month} {year}',
            'date': f'«{day}» {month} {year}',
            'day': day,
            'month': month,
            'year': year,
            'website': 'almazgeobur.ru',
            'company_name_ru': 'АЛМАЗГЕОБУР',
            'company_name_en': 'ALMAZGEOBUR'
        }
        
        # Заменяем все переменные в формате {{ variable }}
        replace_all_variables(ws, row_offset, col_offset, STICKER_ROWS, STICKER_COLS, variables)
        
        # Также заменяем по тексту (для обратной совместимости)
        # Название номенклатуры
        find_and_replace_cell(ws, row_offset, col_offset, STICKER_ROWS, STICKER_COLS, 
                            'nomenclature_name', nom_name)
        find_and_replace_cell(ws, row_offset, col_offset, STICKER_ROWS, STICKER_COLS, 
                            'Коронка', nom_name)
        find_and_replace_cell(ws, row_offset, col_offset, STICKER_ROWS, STICKER_COLS, 
                            'ALFA', nom_name)
        
        # Включаем перенос строк в ячейках с названием номенклатуры (чтобы перенос после ALFA отображался)
        for _r in range(row_offset, row_offset + STICKER_ROWS):
            for _c in range(col_offset, col_offset + STICKER_COLS):
                _cell = ws.cell(_r, _c)
                if isinstance(_cell, MergedCell):
                    continue
                if _cell.value and isinstance(_cell.value, str) and '\n' in _cell.value:
                    _al = _cell.alignment
                    _cell.alignment = Alignment(
                        horizontal=_al.horizontal if _al else 'left',
                        vertical=_al.vertical if _al else 'top',
                        wrap_text=True,
                        indent=getattr(_al, 'indent', 0) if _al else 0,
                        text_rotation=getattr(_al, 'text_rotation', 0) if _al else 0
                    )
        
        # Артикул - ищем ячейку со значением артикула (после "Артикул:" или переменной)
        for row_idx in range(row_offset, row_offset + STICKER_ROWS):
            for col_idx in range(col_offset, col_offset + STICKER_COLS):
                cell = ws.cell(row_idx, col_idx)
                if cell.value and isinstance(cell.value, str):
                    if 'Артикул:' in cell.value or '{{ article }}' in cell.value or '{{article}}' in cell.value:
                        # Если в этой же ячейке есть значение артикула, заменяем его
                        if '{{ article }}' in cell.value or '{{article}}' in cell.value:
                            import re
                            cell.value = re.sub(r'\{\{\s*article\s*\}\}', stock_code, cell.value, flags=re.IGNORECASE)
                        # Следующая ячейка справа - значение артикула
                        value_cell = ws.cell(row_idx, col_idx + 1)
                        if value_cell and not isinstance(value_cell, MergedCell):
                            if value_cell.value is None or value_cell.value == '' or '{{' in str(value_cell.value):
                                value_cell.value = stock_code
                    break
        
        # Высота матрицы
        for row_idx in range(row_offset, row_offset + STICKER_ROWS):
            for col_idx in range(col_offset, col_offset + STICKER_COLS):
                cell = ws.cell(row_idx, col_idx)
                if cell.value and isinstance(cell.value, str) and 'Высота матрицы:' in cell.value:
                    value_cell = ws.cell(row_idx, col_idx + 1)
                    if value_cell and not isinstance(value_cell, MergedCell):
                        value_cell.value = f"{height_val} мм."
                    break
        
        # Промывочные отверстия
        for row_idx in range(row_offset, row_offset + STICKER_ROWS):
            for col_idx in range(col_offset, col_offset + STICKER_COLS):
                cell = ws.cell(row_idx, col_idx)
                if cell.value and isinstance(cell.value, str) and 'Промывочные отверстия:' in cell.value:
                    value_cell = ws.cell(row_idx, col_idx + 1)
                    if value_cell and not isinstance(value_cell, MergedCell):
                        value_cell.value = f"{waterways_val} мм."
                    break
        
        # Типоразмер
        for row_idx in range(row_offset, row_offset + STICKER_ROWS):
            for col_idx in range(col_offset, col_offset + STICKER_COLS):
                cell = ws.cell(row_idx, col_idx)
                if cell.value and isinstance(cell.value, str) and 'Типоразмер:' in cell.value:
                    value_cell = ws.cell(row_idx, col_idx + 1)
                    if value_cell and not isinstance(value_cell, MergedCell):
                        value_cell.value = matrix_val
                    break
        
        # Серийный номер
        find_and_replace_cell(ws, row_offset, col_offset, STICKER_ROWS, STICKER_COLS,
                            'serial_number', serial_number)
        find_and_replace_cell(ws, row_offset, col_offset, STICKER_ROWS, STICKER_COLS,
                            'Серийный номер:', f"Серийный номер: {serial_number}")
        
        # Текст в 7-й строке наклейки — смещение к верху ячейки. Для объединённых ячеек выравнивание
        # задаётся в верхней левой ячейке объединения, иначе у 2-й наклейки текст/штрихкод плывут.
        row_7 = row_offset + 6  # 7-я строка наклейки (1-based)
        if row_7 < row_offset + STICKER_ROWS:
            seen_origins = set()  # (row, col) — уже выставили vertical=top
            for _c in range(col_offset, col_offset + STICKER_COLS):
                _cell = ws.cell(row_7, _c)
                bounds = _find_merged_bounds(ws, row_7, _c)
                if bounds:
                    min_col, min_row, max_col, max_row = bounds
                    origin_row, origin_col = min_row, min_col
                else:
                    origin_row, origin_col = row_7, _c
                origin_key = (origin_row, origin_col)
                if origin_key in seen_origins:
                    continue
                seen_origins.add(origin_key)
                origin_cell = ws.cell(origin_row, origin_col)
                _al = origin_cell.alignment
                origin_cell.alignment = Alignment(
                    horizontal=_al.horizontal if _al else 'center',
                    vertical='top',
                    wrap_text=getattr(_al, 'wrap_text', True) if _al else True,
                    indent=getattr(_al, 'indent', 0) if _al else 0,
                    text_rotation=getattr(_al, 'text_rotation', 0) if _al else 0
                )
        
        # Дата изготовления
        find_and_replace_cell(ws, row_offset, col_offset, STICKER_ROWS, STICKER_COLS,
                            'production_date', f'«{day}» {month} {year}')
        find_and_replace_cell(ws, row_offset, col_offset, STICKER_ROWS, STICKER_COLS,
                            'Дата изготовления:', f'Дата изготовления: «{day}» {month} {year}')
        
        # Усиливаем границы наклейки (делаем их жирными)
        thick_side = Side(style='thick', color='000000')
        thin_side = Side(style='thin', color='000000')
        
        for row_idx in range(row_offset, row_offset + STICKER_ROWS):
            for col_idx in range(col_offset, col_offset + STICKER_COLS):
                cell = ws.cell(row_idx, col_idx)
                
                # Определяем внешние границы
                is_top = (row_idx == row_offset)
                is_bottom = (row_idx == row_offset + STICKER_ROWS - 1)
                is_left = (col_idx == col_offset)
                is_right = (col_idx == col_offset + STICKER_COLS - 1)
                
                # Создаем границы: внешние - жирные, внутренние - тонкие
                cell.border = Border(
                    left=thick_side if is_left else thin_side,
                    right=thick_side if is_right else thin_side,
                    top=thick_side if is_top else thin_side,
                    bottom=thick_side if is_bottom else thin_side
                )
        
        # Вставляем логотип ТОЛЬКО ОДИН РАЗ в объединенную ячейку A (первая колонка)
        if logo_path and os.path.exists(logo_path):
            try:
                from PIL import Image
                import shutil
                
                # Поворачиваем логотип на 90° по часовой (как на шаблоне — текст снизу вверх)
                img = Image.open(logo_path)
                rotated_img = img.rotate(-90, expand=True)
                
                # Сохраняем во временный файл
                temp_logo_dir = '/tmp/agb_stickers'
                os.makedirs(temp_logo_dir, exist_ok=True)
                temp_logo_filename = f'logo_{passport.id}_{sticker_idx}.png'
                temp_logo_path_name = os.path.join(temp_logo_dir, temp_logo_filename)
                
                rotated_img.save(temp_logo_path_name, 'PNG')
                
                if not os.path.exists(temp_logo_path_name):
                    raise FileNotFoundError(f"Не удалось создать временный файл логотипа: {temp_logo_path_name}")
                
                temp_files_to_cleanup.append(temp_logo_path_name)
                
                # Логотип: размер 1.07 см x 3.61 см, посередине ячейки
                logo_img = OpenpyxlImage(temp_logo_path_name)
                # Делаем логотип в 2 раза больше
                logo_img.width = int(LOGO_WIDTH_CM * CM_TO_PX * 2)
                logo_img.height = int(LOGO_HEIGHT_CM * CM_TO_PX * 2)
                
                logo_cell = f"{get_column_letter(col_offset)}{row_offset}"
                _remove_images_in_cell(ws, row_offset, col_offset)
                _add_image_centered(ws, logo_img, row_offset, col_offset)
                print(f"    ✅ Логотип 1.07×3.61 см в {logo_cell}")
            except Exception as e:
                print(f"    ⚠️ Ошибка добавления логотипа: {e}")
                import traceback
                traceback.print_exc()
        
        # Генерируем и вставляем штрихкоды ТОЛЬКО ОДИН РАЗ
        # Штрихкод артикула
        try:
            stock_code_barcode_temp = generate_barcode_image(stock_code, width_mm=40, height_mm=10)
            if stock_code_barcode_temp and os.path.exists(stock_code_barcode_temp):
                temp_barcode_dir = '/tmp/agb_stickers'
                os.makedirs(temp_barcode_dir, exist_ok=True)
                stock_code_barcode_path = os.path.join(temp_barcode_dir, f'barcode_stock_{passport.id}_{sticker_idx}.png')
                import shutil
                shutil.copy2(stock_code_barcode_temp, stock_code_barcode_path)
                temp_files_to_cleanup.append(stock_code_barcode_path)
                
                try:
                    os.unlink(stock_code_barcode_temp)
                except:
                    pass
                
                # Штрихкод номенклатуры (артикул) — посередине: плейсхолдер {{ stock_code }} или 1-я строка, колонка B (середина)
                barcode_row, barcode_col = _find_cell_with_placeholder(ws, row_offset, col_offset, STICKER_ROWS, STICKER_COLS, 'stock_code')
                if barcode_row is None:
                    # По умолчанию: 1-я строка, средняя колонка (B при A,B,C)
                    barcode_row = row_offset
                    barcode_col = col_offset + (STICKER_COLS - 1) // 2 if STICKER_COLS >= 2 else col_offset + 1
                _remove_images_in_cell(ws, barcode_row, barcode_col)
                _cell = ws.cell(barcode_row, barcode_col)
                if not isinstance(_cell, MergedCell) and _cell.value and '{{' in str(_cell.value):
                    import re
                    _cell.value = re.sub(r'\{\{\s*stock_code\s*\}\}', '', str(_cell.value), flags=re.IGNORECASE).strip() or None
                if not isinstance(_cell, MergedCell):
                    _cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=getattr(_cell.alignment, 'wrap_text', False) if _cell.alignment else False)
                barcode_img = OpenpyxlImage(stock_code_barcode_path)
                # Под ширину колонки B=200px: делаем штрихкод крупнее и визуально ровнее
                barcode_img.height = 40
                barcode_img.width = 180
                _add_image_centered(ws, barcode_img, barcode_row, barcode_col)
                print(f"    ✅ Штрихкод номенклатуры (артикул) по центру в {get_column_letter(barcode_col)}{barcode_row}")
        except Exception as e:
            print(f"    ⚠️ Ошибка добавления штрихкода артикула: {e}")
            import traceback
            traceback.print_exc()
        
        # Штрихкод серийного номера
        try:
            serial_number_barcode_temp = generate_barcode_image(serial_number, width_mm=40, height_mm=10)
            if serial_number_barcode_temp and os.path.exists(serial_number_barcode_temp):
                temp_barcode_dir = '/tmp/agb_stickers'
                os.makedirs(temp_barcode_dir, exist_ok=True)
                serial_number_barcode_path = os.path.join(temp_barcode_dir, f'barcode_serial_{passport.id}_{sticker_idx}.png')
                import shutil
                shutil.copy2(serial_number_barcode_temp, serial_number_barcode_path)
                temp_files_to_cleanup.append(serial_number_barcode_path)
                
                try:
                    os.unlink(serial_number_barcode_temp)
                except:
                    pass
                
                # 1. Находим строку с текстом "Серийный номер" в пределах этой наклейки
                label_row = None
                for r in range(row_offset, row_offset + STICKER_ROWS):
                    for c in range(col_offset, col_offset + STICKER_COLS):
                        val = _cell_display_value(ws, r, c)
                        if val and isinstance(val, str) and 'Серийный номер' in val:
                            label_row = r
                            break
                    if label_row is not None:
                        break
                
                # 2. Штрихкод ставим в ЭТУ ЖЕ СТРОКУ (не под датой), в средней колонке
                if label_row is None:
                    label_row = row_7  # fallback: 7-я строка наклейки
                barcode_row = label_row
                barcode_col = col_offset + (STICKER_COLS - 1) // 2 if STICKER_COLS >= 2 else col_offset + 1
                
                _remove_images_in_cell(ws, barcode_row, barcode_col)
                _cell = ws.cell(barcode_row, barcode_col)
                if not isinstance(_cell, MergedCell) and _cell.value and '{{' in str(_cell.value):
                    import re
                    _cell.value = re.sub(r'\{\{\s*serial_number_code\s*\}\}', '', str(_cell.value), flags=re.IGNORECASE).strip() or None
                # Выравнивание ячейки не трогаем — текст «Серийный номер: XXX» уже по верху (row_7),
                # штрихкод будет добавлен по центру ячейки (как у первой наклейки)
                barcode_img = OpenpyxlImage(serial_number_barcode_path)
                barcode_img.height = 40
                barcode_img.width = 180
                # Штрихкод по центру ячейки (горизонтально и вертикально) — аналогично первой наклейке
                _add_image_centered(ws, barcode_img, barcode_row, barcode_col)
                barcode_cell = f"{get_column_letter(barcode_col)}{barcode_row}"
                print(f"    ✅ Серийный номер: текст сверху, штрихкод по центру ячейки — {barcode_cell}")
        except Exception as e:
            print(f"    ⚠️ Ошибка добавления штрихкода серийного номера: {e}")
            import traceback
            traceback.print_exc()

        # Увеличиваем счётчик и после каждой второй наклейки ставим разрыв страницы,
        # чтобы при печати на листе было по 2 наклейки.
        stickers_generated += 1
        if stickers_generated % 2 == 0:
            break_row = row_offset + STICKER_ROWS - 1
            try:
                ws.row_breaks.append(Break(id=break_row))
                print(f"    📄 Установлен разрыв страницы после строки {break_row} (наклейки {stickers_generated})")
            except Exception as e:
                print(f"    ⚠️ Не удалось установить разрыв страницы после строки {break_row}: {e}")

    # Сохраняем в память
    buffer = io.BytesIO()
    try:
        wb.save(buffer)
        buffer.seek(0)
        excel_content = buffer.getvalue()
    finally:
        # Удаляем временные файлы ПОСЛЕ сохранения
        for temp_file in temp_files_to_cleanup:
            try:
                if os.path.exists(temp_file):
                    os.unlink(temp_file)
            except Exception as cleanup_err:
                print(f"    ⚠️ Не удалось удалить временный файл {temp_file}: {cleanup_err}")
    
    print(f"✅ Excel файл с наклейками успешно сгенерирован, размер: {len(excel_content)} байт")
    sys.stdout.flush()
    
    return excel_content
